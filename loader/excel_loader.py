"""
Excel 探索用例加载器

读取 Excel 中的探索测试用例，解析为标准格式送入探索模式执行。
支持 {{VAR_NAME}} 环境变量替换。

列包含：用例编号、用例名称、前置条件、步骤描述、预期结果
"""

import os
import re

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

_VAR_PATTERN = re.compile(r'\{\{(\w+)\}\}')


def _resolve_env_vars(text: str) -> str:
    if not text or "{{" not in text:
        return text

    def _replacer(match):
        var_name = match.group(1)
        value = os.environ.get(var_name)
        if value is None:
            return match.group(0)
        return value

    return _VAR_PATTERN.sub(_replacer, text)


# ── 列名映射（支持中文/英文列名） ──
EXPECTED_COLUMNS = {
    "case_id":     ["用例编号", "case_id", "id"],
    "case_name":   ["用例名称", "case_name", "name"],
    "preconditions": ["前置条件", "preconditions"],
    "steps_desc":  ["步骤描述", "steps_desc", "step_desc"],
    "expected":    ["预期结果", "expected", "expected_result"],
    "module":      ["模块", "module"],
    "start_url":   ["起始网址", "start_url", "url"],
}


def _find_column(col_name: str, headers: list) -> int:
    """在 headers 中查找目标列（支持中文/英文别名），返回 0-based 索引"""
    aliases = EXPECTED_COLUMNS.get(col_name, [col_name])
    for i, h in enumerate(headers):
        h_stripped = str(h or "").strip().lower()
        for alias in aliases:
            if h_stripped == alias.lower():
                return i
    return -1


def load_excel_cases(filepath: str) -> list[dict]:
    """
    读取 Excel 探索用例，返回统一格式的用例列表。

    返回字段：
        case_id, case_name, preconditions, steps_desc, expected,
        module, start_url, _steps_parsed (步骤goal列表)
    """
    if openpyxl is None:
        raise ImportError("需要 openpyxl 库: pip install openpyxl")
    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"Excel 文件不存在: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # 按别名映射获取列索引
    col_idx = {}
    for key in EXPECTED_COLUMNS:
        col_idx[key] = _find_column(key, header_row)

    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue

        row_list = [str(v or "").strip() if v is not None else "" for v in row]

        def _cell(key):
            idx = col_idx.get(key, -1)
            if 0 <= idx < len(row_list):
                return _resolve_env_vars(row_list[idx])
            return ""

        case_id = _cell("case_id")
        if not case_id:
            continue

        steps_desc = _cell("steps_desc")
        steps_parsed = [s.strip() for s in steps_desc.replace("\n", "|").split("|") if s.strip()]

        case = {
            "case_id": case_id,
            "case_name": _cell("case_name") or case_id,
            "preconditions": _cell("preconditions"),
            "steps_desc": steps_desc,
            "expected": _cell("expected"),
            "module": _cell("module"),
            "start_url": _cell("start_url"),
            "_steps_parsed": steps_parsed,
        }
        cases.append(case)

    wb.close()
    return cases


# ── 命令行预览 ──
if __name__ == "__main__":
    import sys
    target = sys.argv[1] if len(sys.argv) > 1 else "test_cases/explore_cases.xlsx"
    cases = load_excel_cases(target)
    print(f"发现 {len(cases)} 个探索用例:")
    for c in cases:
        steps = c["_steps_parsed"]
        print(f"  {c['case_id']:8s} {c['case_name']:20s} | {len(steps)} 步 | 模块: {c.get('module','-')}")
        for s in steps:
            print(f"      → {s}")
