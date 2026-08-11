"""Deterministic, fail-closed compiler for the Web-agent curriculum workbook."""

from __future__ import annotations

from dataclasses import dataclass, field
from hashlib import sha256
from pathlib import Path
import os
import re

import openpyxl

from core.credential_vault import PII_EMAIL_PATTERN

SHEET_NAME = "自动化测试用例"
HEADERS = ("用例ID", "起始网址", "模块", "测试场景", "优先级", "前置条件", "操作步骤", "期望结果", "是否执行")
PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2, "P3": 3}
_DESTRUCTIVE = ("删除", "注销", "销毁")
_MUTATING = ("新增", "添加", "编辑", "修改", "保存", "上传", "更改", "更新", "备注", "创建")
_NEGATIVE_AUTH = ("错误账号", "错误密码", "账号为空", "密码为空", "空账号", "空密码", "不输入", "未输入", "登录失败")
_VAGUE_DATA = ("存在的用户", "不存在的手机号", "存在数据的时间范围", "选择一个顾客", "任意顾客")
_VAGUE_EXPECTED = {"搜索结果", "显示正确", "操作成功", "筛选成功"}
_PHONE_LITERAL = re.compile(r"(?<!\d)1[3-9]\d{9}(?!\d)")
_NUMBER_PREFIX = re.compile(r"^\s*\d+\s*[.．、)]\s*")
_INLINE_NUMBER = re.compile(r"(?<!^)\s+(?=\d+\s*[.．、)])")


@dataclass(frozen=True)
class CurriculumIssue:
    code: str
    message: str
    case_id: str = ""
    row: int = 0

    def to_dict(self) -> dict:
        return self.__dict__.copy()


@dataclass(frozen=True)
class CurriculumCase:
    case_id: str
    start_url: str
    module: str
    case_name: str
    priority: str
    preconditions: str
    steps: tuple[str, ...]
    expected: str
    enabled: bool
    risk: str
    row: int
    warnings: tuple[str, ...] = ()
    capability_gaps: tuple[str, ...] = ()

    @property
    def blocked(self) -> bool:
        return (
            bool(self.capability_gaps)
            or not self.start_url
            or not self.steps
            or not self.expected
        )

    def runner_steps(self) -> list[dict]:
        goals = []
        if "登录" in self.preconditions and self.module != "账号登录":
            goals.extend((
                "在账号输入框输入 {{credential.username}}",
                "在密码输入框输入 {{credential.password}}",
                "选择配置的门店",
                "点击登录按钮并确认进入登录后的首页",
            ))
        if "顾客列表" in self.preconditions:
            goals.append("点击顾客档案菜单")
        source_steps = list(self.steps)
        if (
            "顾客列表" in self.preconditions
            and source_steps
            and "进入顾客档案列表" in source_steps[0]
        ):
            source_steps = source_steps[1:]
        goals.extend(_credential_goal(goal, self.risk) for goal in source_steps)
        if self.risk == "negative_auth" and any(
            marker in self.case_name for marker in ("错误账号", "错误密码")
        ):
            invalid_goal = next(
                goal for goal in goals if "{{credential.invalid." in goal
            )
            click_goal = next(
                goal for goal in goals if "点击" in goal and "登录" in goal
            )
            goals = [
                "输入 {{credential.username}}",
                "输入 {{credential.password}}",
                "选择门店",
                invalid_goal,
                click_goal,
            ]
        result = [{"goal": goal} for goal in goals]
        if result:
            result[-1]["success_criteria"] = compile_acceptance(self.expected)
        return result

    def evaluation_case(self) -> dict:
        return {"case_id": self.case_id, "case_name": self.case_name, "module": self.module}

    def to_dict(self) -> dict:
        result = self.__dict__.copy()
        result["steps"] = list(self.steps)
        result["warnings"] = list(self.warnings)
        result["capability_gaps"] = list(self.capability_gaps)
        return result


@dataclass
class CurriculumCatalog:
    source_path: str
    source_hash: str
    cases: list[CurriculumCase] = field(default_factory=list)
    errors: list[CurriculumIssue] = field(default_factory=list)
    warnings: list[CurriculumIssue] = field(default_factory=list)

    def summary(self) -> dict:
        priorities, risks = {}, {}
        for case in self.cases:
            priorities[case.priority] = priorities.get(case.priority, 0) + 1
            risks[case.risk] = risks.get(case.risk, 0) + 1
        return {
            "source": self.source_path,
            "source_hash": self.source_hash,
            "total": len(self.cases),
            "enabled": sum(case.enabled for case in self.cases),
            "disabled": sum(not case.enabled for case in self.cases),
            "blocked": sum(case.blocked for case in self.cases),
            "priorities": dict(sorted(priorities.items())),
            "risks": dict(sorted(risks.items())),
            "errors": len(self.errors),
            "warnings": len(self.warnings),
        }


def split_steps(text: str) -> tuple[str, ...]:
    normalized = str(text or "").replace("\\n", "\n").replace("\r", "\n")
    normalized = _INLINE_NUMBER.sub("\n", normalized)
    return tuple(
        cleaned
        for part in re.split(r"\n+|\|", normalized)
        if (cleaned := _NUMBER_PREFIX.sub("", part).strip())
    )


def compile_acceptance(expected: str) -> str | dict:
    text = str(expected or "").strip()
    display = re.match(r"^(?:展示|显示|提示|红色提示)[：:]\s*(.+)$", text)
    if display:
        values = [
            item.strip()
            for item in re.split(r"[、，,|]", display.group(1))
            if item.strip()
        ]
        if values:
            return {"text_contains": values}
    if not text.startswith("登录成功"):
        url_contains = re.search(
            r"URL\s*包含\s*(/[A-Za-z0-9_./-]+)", text, re.I
        )
        if url_contains:
            return {"url_contains": [url_contains.group(1)]}
        navigation = re.match(r"^跳转到(.+?)(?:页面|页)?$", text)
        if navigation:
            target = navigation.group(1).strip()
            if target.endswith("列表"):
                target = target[:-2]
            return {"url_changed": True, "text_contains": [target]}
        entered = re.match(r"^跳转进入(.+?)(?:界面|页面|页)$", text)
        if entered:
            return {
                "url_changed": True,
                "text_contains": [entered.group(1).strip()],
            }
    return text


def _priority(value: str) -> tuple[str, str]:
    raw = str(value or "").strip()
    upper = raw.upper()
    if upper in PRIORITY_ORDER:
        return upper, "" if raw == upper else f"优先级 {raw} 已规范为 {upper}"
    if raw in {"中", "中等", "一般"}:
        return "P2", f"优先级 {raw} 已规范为 P2"
    return "P3", f"未知优先级 {raw or '<空>'} 已降级为 P3"


def _enabled(value: str) -> bool:
    return str(value or "").strip().lower() in {"是", "yes", "y", "true", "1"}


def _risk(module: str, case_name: str, steps: tuple[str, ...]) -> str:
    text = " ".join((module, case_name, *steps))
    if any(word in text for word in _DESTRUCTIVE):
        return "destructive"
    if module == "账号登录" and any(word in text for word in _NEGATIVE_AUTH):
        return "negative_auth"
    if any(word in text for word in _MUTATING):
        return "mutation"
    return "read_only"


def _credential_goal(goal: str, risk: str) -> str:
    result = str(goal)
    result = re.sub(
        r"输入错误账号(?:为)?[A-Za-z0-9_.@+-]*",
        "输入 {{credential.invalid.username}}",
        result,
    )
    result = re.sub(
        r"输入错误密码(?:为)?[A-Za-z0-9_.@+-]*",
        "输入 {{credential.invalid.password}}",
        result,
    )
    for source, target in (
        ("输入正确账号", "输入 {{credential.username}}"),
        ("输入正确密码", "输入 {{credential.password}}"),
    ):
        result = result.replace(source, target)
    if risk != "negative_auth":
        result = result.replace("输入账号", "输入 {{credential.username}}")
        result = result.replace("输入密码", "输入 {{credential.password}}")
    return result


def load_curriculum(filepath: str | os.PathLike, default_start_url: str = "") -> CurriculumCatalog:
    path = Path(filepath)
    if not path.is_file():
        raise FileNotFoundError(f"课程表不存在: {path}")
    catalog = CurriculumCatalog(str(path), sha256(path.read_bytes()).hexdigest())
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"缺少工作表: {SHEET_NAME}")
        rows = workbook[SHEET_NAME].iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(rows)]
        missing = [name for name in HEADERS if name not in headers]
        if missing:
            raise ValueError("缺少必需列: " + ", ".join(missing))
        columns = {name: headers.index(name) for name in HEADERS}
        last_url = str(default_start_url or "").strip()
        seen_ids, seen_scenarios = set(), set()
        for row_number, values in enumerate(rows, 2):
            cells = [str(value or "").strip() for value in values]
            if not any(cells):
                continue

            def get(name: str) -> str:
                index = columns[name]
                return cells[index] if index < len(cells) else ""

            case_id = get("用例ID")
            if not case_id:
                catalog.errors.append(CurriculumIssue(
                    "missing_case_id",
                    "用例ID不能为空",
                    row=row_number,
                ))
                continue
            explicit_url = get("起始网址")
            if explicit_url:
                last_url = explicit_url
            start_url = explicit_url or last_url or str(default_start_url or "").strip()
            module, case_name = get("模块"), get("测试场景") or case_id
            steps, expected = split_steps(get("操作步骤")), get("期望结果")
            priority, priority_warning = _priority(get("优先级"))
            warnings = [priority_warning] if priority_warning else []
            gaps = []
            for condition, code, message in (
                (not start_url, "missing_start_url", "没有表内 URL，也没有 LOGIN_URL 兜底"),
                (not steps, "missing_steps", "操作步骤不能为空"),
                (not expected, "missing_expected", "期望结果不能为空"),
                (case_id in seen_ids, "duplicate_case_id", "用例ID重复"),
            ):
                if condition:
                    catalog.errors.append(CurriculumIssue(code, message, case_id, row_number))
            seen_ids.add(case_id)
            scenario_key = (module, case_name)
            if scenario_key in seen_scenarios:
                warnings.append("同一模块存在重复测试场景")
            seen_scenarios.add(scenario_key)
            combined = " ".join((
                case_id, module, case_name, get("前置条件"), *steps, expected,
            ))
            risk = _risk(module, case_name, steps)
            if any(token in combined for token in _VAGUE_DATA):
                gaps.append("缺少可复现的具名测试数据")
            if expected in _VAGUE_EXPECTED:
                gaps.append("期望结果过于模糊，无法形成证据化断言")
            if _PHONE_LITERAL.search(combined):
                gaps.append("用例含明文手机号，需改为脱敏 fixture 引用")
            if PII_EMAIL_PATTERN.search(combined):
                gaps.append("用例含明文邮箱，需改为脱敏 fixture 引用")
            if expected.startswith("卡片包含"):
                gaps.append("卡片字段需要结构化 DOM oracle，不能用全文文本断言")
            if risk == "negative_auth":
                gaps.append("负向登录需显式测试凭证/空值协议，默认禁止执行")
            if "顾客详情页" in get("前置条件"):
                gaps.append("顾客详情前置条件缺少专用测试顾客 fixture")
            if risk in {"mutation", "destructive"} and "恢复" not in combined:
                warnings.append("写操作没有明确恢复步骤")
            enabled = _enabled(get("是否执行"))
            if not enabled:
                warnings.append("源表标记为不执行")
            case = CurriculumCase(
                case_id, start_url, module, case_name, priority,
                get("前置条件"), steps, expected, enabled, risk,
                row_number, tuple(warnings), tuple(gaps),
            )
            catalog.cases.append(case)
            catalog.warnings.extend(
                CurriculumIssue("case_warning", item, case_id, row_number)
                for item in (*warnings, *gaps)
            )
    finally:
        workbook.close()
    return catalog


__all__ = [
    "CurriculumCase", "CurriculumCatalog", "CurriculumIssue",
    "PRIORITY_ORDER", "compile_acceptance", "load_curriculum", "split_steps",
]