"""
Excel 测试用例加载器

从 xlsx 读取「用例清单」和「步骤明细」两个 Sheet，
合并输出与 JSON 文件完全兼容的 test case dict 列表，
可直接替换 run_batch.py 中的 load_test_cases()。

Excel 模板字段说明：
  Sheet1「用例清单」: case_id, case_name, module, mode, start_url, task, priority, tags, enabled, description, login_required, credential_key
  Sheet2「步骤明细」: case_id, step_no, goal, assert_type, assert_target, assert_value, remark

变量替换：
  所有字符串字段支持 {{VAR_NAME}} 语法，加载时自动从 .env 文件替换。
  推荐在 task / goal / start_url 中使用，例如：
    task: 打开{{LOGIN_URL}}，使用{{LOGIN_USERNAME}}登录
"""

import os
import re
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("[excel_loader] 请安装 openpyxl: pip install openpyxl")

# 加载 .env 到环境变量（dotenv 在 settings.py 已加载，但 excel_loader 可能独立使用）
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ── 变量替换 ──

_VAR_PATTERN = re.compile(r'\{\{(\w+)\}\}')


def _resolve_env_vars(text: str) -> str:
    """
    将字符串中的 {{VAR_NAME}} 替换为环境变量值。
    未设置的变量保持原样并打印警告。
    """
    if not text or "{{" not in text:
        return text

    def _replacer(match):
        var_name = match.group(1)
        value = os.environ.get(var_name)
        if value is None:
            print(f"  [excel_loader] [WARN] 环境变量 {match.group(0)} 未设置，保留原样")
            return match.group(0)
        return value

    return _VAR_PATTERN.sub(_replacer, text)


# ── 字段列表 ──

CASE_HEADERS = [
    "case_id",          # 用例编号（必填）
    "case_name",        # 用例名称
    "module",           # 所属模块
    "mode",             # auto / steps
    "start_url",        # 起始网址
    "task",             # auto 模式的任务描述
    "priority",         # P0-P3
    "tags",             # 标签
    "enabled",          # Y/N
    "description",      # 描述
    "login_required",   # Y/N 是否需要预先登录
    "credential_key",   # 凭证组名（空=默认 LOGIN_USERNAME/PASSWORD）
]

STEP_HEADERS = [
    "case_id",          # 归属用例编号
    "step_no",          # 步骤序号
    "goal",             # 步骤目标
    "assert_type",      # 断言类型
    "assert_target",    # 断言目标值
    "assert_value",     # 断言补充值
    "remark",           # 备注
]


# ── 主加载函数 ──

def load_from_xlsx(
    filepath: str,
    only_enabled: bool = True,
) -> list[dict]:
    """
    从 xlsx 加载全部测试用例，返回 list[dict]。

    参数:
        filepath: xlsx 文件路径
        only_enabled: 仅加载 enabled=Y 的用例（默认 True）

    返回:
        与 load_test_cases() 完全兼容的 dict 列表
    """
    if openpyxl is None:
        raise ImportError("需要 openpyxl 库: pip install openpyxl")

    if not os.path.isfile(filepath):
        raise FileNotFoundError(f"Excel 文件不存在: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── 1. 读取用例清单 ──
    if "用例清单" not in wb.sheetnames:
        raise ValueError(f'缺少「用例清单」Sheet，请使用模板生成')
    ws1 = wb["用例清单"]
    rows1 = list(ws1.iter_rows(min_row=2, values_only=True))
    headers1 = [str(cell.value or "") for cell in next(ws1.iter_rows(min_row=1, max_row=1))]

    cases = []
    for row in rows1:
        if all(v is None for v in row):
            continue
        row_dict = dict(zip(headers1, row))

        case_id = str(row_dict.get("case_id") or "").strip()
        if not case_id:
            continue

        # 跳过禁用
        enabled = str(row_dict.get("enabled") or "Y").strip().upper()
        if only_enabled and enabled == "N":
            continue

        login_required = str(row_dict.get("login_required") or "N").strip().upper()
        credential_key = str(row_dict.get("credential_key") or "").strip()

        start_url = _resolve_env_vars(str(row_dict.get("start_url") or "").strip())
        # login_required=Y 且未填 start_url 时，尝试用 LOGIN_URL
        if not start_url and login_required == "Y":
            start_url = _resolve_env_vars("{{LOGIN_URL}}")

        case: dict = {
            "name": _resolve_env_vars(str(row_dict.get("case_name") or case_id)).strip(),
            "start_url": start_url or "https://www.baidu.com",
            "mode": str(row_dict.get("mode") or "auto").strip().lower(),
            "_case_id": case_id,
            "_module": _resolve_env_vars(str(row_dict.get("module") or "").strip()),
            "_priority": str(row_dict.get("priority") or "").strip(),
            "_tags": str(row_dict.get("tags") or "").strip(),
            "_enabled": enabled,
            "_description": _resolve_env_vars(str(row_dict.get("description") or "").strip()),
            "_login_required": login_required == "Y",
            "_credential_key": credential_key,
        }

        if case["mode"] == "auto":
            task_raw = str(row_dict.get("task") or case["name"]).strip()
            case["task"] = _resolve_env_vars(task_raw)
        elif case["mode"] == "steps":
            case["steps"] = []  # 后面从步骤明细填充

        cases.append(case)

    # ── 2. 读取步骤明细，合并到 steps 模式用例 ──
    if "步骤明细" in wb.sheetnames:
        ws2 = wb["步骤明细"]
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        headers2 = [str(cell.value or "") for cell in next(ws2.iter_rows(min_row=1, max_row=1))]

        steps_by_case = {}
        for row in rows2:
            if all(v is None for v in row):
                continue
            row_dict = dict(zip(headers2, row))

            cid = str(row_dict.get("case_id") or "").strip()
            if not cid:
                continue

            goal_raw = str(row_dict.get("goal") or "").strip()
            if not goal_raw:
                continue

            goal = _resolve_env_vars(goal_raw)

            step: dict = {
                "step": int(row_dict.get("step_no") or 1),
                "goal": goal,
            }

            # 断言字段也支持变量替换
            at = _resolve_env_vars(str(row_dict.get("assert_type") or "").strip())
            ta = _resolve_env_vars(str(row_dict.get("assert_target") or "").strip())
            av = _resolve_env_vars(str(row_dict.get("assert_value") or "").strip())

            assert_expr = _build_assert_expr(at, ta, av)
            if assert_expr:
                step["assert"] = assert_expr

            validation = _build_validation(at, ta, av)
            if validation:
                step["validation"] = validation

            steps_by_case.setdefault(cid, []).append(step)

        # 合并到对应 case
        for case in cases:
            cid = case["_case_id"]
            if cid in steps_by_case and case["mode"] == "steps":
                case["steps"] = sorted(steps_by_case[cid], key=lambda s: s["step"])

    wb.close()
    return cases


# ── 断言构建 ──

def _build_assert_expr(atype: str, target: str, val: str) -> Optional[str]:
    """
    将 assert_type / assert_target 合并为兼容的 assert 字符串。
    例如: url_contains + baidu → "URL包含baidu"
    """
    if val:
        target = val
    if not atype or not target:
        return None
    mapping = {
        "url_contains": "URL包含",
        "url_exact": "URL等于",
        "title_contains": "标题包含",
        "text_exists": "页面包含文字",
        "element_visible": "元素可见",
    }
    prefix = mapping.get(atype, atype)
    return f"{prefix}{target}"


def _build_validation(atype: str, target: str, val: str) -> Optional[dict]:
    """构建结构化验证条件"""
    if not atype or not target:
        return None
    return {
        "type": atype,
        "target": target,
        "value": val or target,
    }


# ── 预览 ──

def list_excel_cases(filepath: str) -> list[dict]:
    """轻量预览：仅读取用例清单。"""
    if not os.path.isfile(filepath):
        return []
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if "用例清单" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["用例清单"]
    headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        d = dict(zip(headers, row))
        cid = str(d.get("case_id") or "").strip()
        if not cid:
            continue
        result.append({
            "case_id": cid,
            "case_name": str(d.get("case_name") or cid).strip(),
            "mode": str(d.get("mode") or "auto").strip().lower(),
            "priority": str(d.get("priority") or "").strip(),
            "enabled": str(d.get("enabled") or "Y").strip().upper(),
            "login_required": str(d.get("login_required") or "N").strip().upper(),
        })
    wb.close()
    return result


# ── 命令行 ──

if __name__ == "__main__":
    import sys
    target = sys.argv[1] if len(sys.argv) > 1 else "test_cases/test_cases.xlsx"

    print(f"[excel_loader] 读取 {target}")
    print()

    preview = list_excel_cases(target)
    if preview:
        print(f"发现 {len(preview)} 个用例:")
        for c in preview:
            status = "启用" if c["enabled"] == "Y" else "禁用"
            login = " [需登录]" if c["login_required"] == "Y" else ""
            print(f"  {c['case_id']:8s} [{status}]{login} {c['case_name']:20s} mode={c['mode']:6s} pri={c['priority']}")
    else:
        print("(无有效用例)")

    print()

    cases = load_from_xlsx(target)
    print(f"load_from_xlsx 返回 {len(cases)} 个用例:")
    for c in cases:
        login_tag = " [需登录]" if c.get("_login_required") else ""
        if c["mode"] == "auto":
            print(f"  {c['_case_id']:8s} auto{login_tag} | {c['name']:20s} | task: {c['task'][:50]}")
        else:
            n_steps = len(c.get("steps", []))
            print(f"  {c['_case_id']:8s} steps{login_tag}| {c['name']:20s} | {n_steps} 个步骤")
            for s in c.get("steps", []):
                a = s.get("assert") or "-"
                print(f"          步骤{s['step']}: {s['goal']:30s} assert: {a}")
