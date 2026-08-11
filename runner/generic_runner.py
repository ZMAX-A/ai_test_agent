"""
通用测试执行引擎 — 数据驱动，一条脚本跑 standard.xlsx 中所有用例

架构（参考 text 项目）:
  StepExecutor — 解析「元素定位器」「操作类型」「输入数据」→ Playwright 操作
  AssertionExecutor — 解析「断言类型」「验证点」→ 断言检查
  前置条件处理 — "已登录"自动登录，"打开登录页"仅导航

用法:
  python -m runner.generic_runner              # 执行全部用例
  python -m runner.generic_runner --case TC001 # 执行单个
"""

import sys, os, re, time, logging
from pathlib import Path
from playwright.sync_api import sync_playwright, Page, TimeoutError as PwTimeout
from openpyxl import load_workbook
from openpyxl.styles import Font

sys.path.insert(0, str(Path(__file__).parent.parent))

from config.settings import settings
from standard.store import StandardCaseStore
from core.credential_vault import CredentialVault

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(name)s] %(message)s")
logger = logging.getLogger("generic_runner")


def _split_steps(value: str) -> list[str]:
    """解析多步骤字段。

    新格式使用 ``|``，避免与 CSS 组合选择器中的逗号冲突；同时兼容旧表的
    逗号分隔格式。
    """
    if not value:
        return []
    delimiter = "|" if "|" in value else ","
    return [part.strip() for part in value.split(delimiter)]


def _precondition_mode(preconditions: str) -> str:
    """把自然语言前置条件归一为稳定的执行模式。"""
    pc = (preconditions or "").strip()
    if not pc:
        return "none"
    if "打开登录" in pc or "未登录" in pc:
        return "open_login"
    if any(keyword in pc for keyword in ("已登录", "登录成功", "需配置", "登录")):
        return "auto_login"
    return "ensure_page"


# ═══════════════════════════════════════════════
#  StepExecutor — 关键字驱动的操作执行器
# ═══════════════════════════════════════════════

class StepExecutor:
    def __init__(self, page: Page, base_url: str = ""):
        self.page = page
        m = re.match(r"(https?://[^/]+)", base_url)
        self.base_url = m.group(1) if m else base_url.rstrip("/")
        self.vault = CredentialVault()

    def execute(self, locators_str: str, operations_str: str, data_str: str):
        if not operations_str or not operations_str.strip():
            return
        locators = _split_steps(locators_str)
        operations = _split_steps(operations_str)
        # 空数据也保留位置，避免多输入框场景发生数据错位。
        data_parts = [d.strip() for d in data_str.split("|")] if data_str else []
        while len(data_parts) < len(operations):
            data_parts.append("")
        data_idx = 0

        for i, op in enumerate(operations):
            if len(operations) == 1 and len(locators) > 1:
                loc = locators_str
            else:
                loc = locators[i] if i < len(locators) else ""
            logger.info(f"  步骤[{i}]: {op} | 定位器={loc[:60]}")

            if op == "input":
                data = data_parts[data_idx]; data_idx += 1
                self._input(loc, data)
            elif op == "click":
                self._click(loc)
            elif op == "select":
                data = data_parts[data_idx]; data_idx += 1
                self._select(loc, data)
            elif op == "verify":
                self._verify(loc)
            elif op == "hover":
                self._hover(loc)
            elif op == "scroll":
                self._scroll(loc)
            elif op == "wait":
                try: seconds = int(loc) if loc else 1
                except ValueError: seconds = 1
                self.page.wait_for_timeout(seconds * 1000)
            elif op == "nav":
                target = data_parts[data_idx] if data_idx < len(data_parts) else loc
                data_idx += 1
                if not target:
                    target = loc
                if not target.startswith("http") and self.base_url:
                    target = self.base_url + target
                self.page.goto(target, wait_until="domcontentloaded", timeout=30000)
                if settings.NAVIGATION_SETTLE_MS > 0:
                    self.page.wait_for_timeout(settings.NAVIGATION_SETTLE_MS)
            elif op == "upload":
                data = data_parts[data_idx]; data_idx += 1
                self._upload(loc, data)
            else:
                raise ValueError(f"不支持的操作类型: {op}")

    def _input(self, locator: str, text: str):
        if not locator: return
        resolved = (
            self.vault.resolve_text(text)
            if "{{credential." in str(text or "") else text
        )
        el = self.page.locator(locator).first
        el.wait_for(state="visible", timeout=10000)
        el.fill(resolved)
        logger.info(f"    输入: {self.vault.sanitize_text(resolved)[:30]}")

    def _click(self, locator: str):
        if not locator: return
        try: self.page.wait_for_load_state("networkidle", timeout=5000)
        except Exception: pass
        self.page.wait_for_timeout(300)
        url_before = self.page.url.rstrip("/")
        logger.info(f"    click前URL: {url_before}")
        locators = [l.strip() for l in locator.split(",")]

        for loc_idx, loc in enumerate(locators):
            if not loc: continue
            try:
                self._click_one(loc, loc_idx, url_before)
            except Exception as e:
                if loc_idx < len(locators) - 1:
                    logger.info(f"    定位器[{loc_idx}]失败: {str(e)[:50]}, 试下一个")
                    continue
                raise
            self.page.wait_for_timeout(500)
            new_url = self.page.url.rstrip("/")
            if new_url != url_before:
                logger.info(f"    ✅ 页面已跳转: {new_url}")
                return
            if loc_idx < len(locators) - 1:
                logger.info(f"    点了但未跳转, 继续试下一个...")

    def _click_one(self, locator: str, idx: int = 0, url_before: str = ""):
        if locator.startswith("text="):
            name = locator[5:]
            self.page.locator(f"text={name}").first.wait_for(state="attached", timeout=3000)
            self.page.wait_for_timeout(300)
            matches = self.page.locator(f"text={name}").all()

            def score(el):
                try:
                    tag = el.evaluate("e => e.tagName.toLowerCase()")
                    role = (el.get_attribute("role") or "").lower()
                    cls = (el.get_attribute("class") or "").lower()
                    parent = el.evaluate("""e => {
                        let p = e.parentElement;
                        if (!p) return '';
                        return p.tagName.toLowerCase() + ' ' + (p.className || '');
                    }""")
                    if tag in ("a", "button") or role in ("link", "button", "menuitem"):
                        return 0
                    if "menu" in cls or "menu" in parent or tag == "li" or parent.startswith("li"):
                        return 1
                    if role or "click" in cls.lower() or "btn" in cls.lower():
                        return 2
                    return 9
                except Exception:
                    return 9

            matches.sort(key=score)
            logger.info(f"    text匹配{len(matches)}个，按交互性排序后逐个试")

            for i, el in enumerate(matches):
                try:
                    tag = el.evaluate("e => e.tagName.toLowerCase()")
                    el.click(timeout=2000)
                    self.page.wait_for_timeout(600)
                    new_url = self.page.url.rstrip("/")
                    logger.info(f"    text[{i}](tag={tag}) url_before={url_before!r} new={new_url!r}")
                    if new_url != url_before:
                        logger.info(f"    ✅ 点击+跳转(text[{i}]): {name}")
                        return
                    logger.info(f"    text[{i}] 未跳转，继续...")
                except Exception:
                    continue
            if matches:
                matches[0].click(force=True, timeout=2000)
            return

        el = self.page.locator(locator).first
        el.wait_for(state="visible", timeout=10000)
        el.click()
        logger.info(f"    点击成功[{idx}]: {locator[:40]}")
        if 'select' in locator.lower() or 'store' in locator.lower():
            self.page.wait_for_timeout(800)
            try:
                opt = self.page.locator('.ant-select-item-option').first
                if opt.is_visible(timeout=1000): opt.click()
            except Exception: pass

    def _select(self, locator: str, option: str):
        if not locator: return
        if locator.startswith("text="):
            name = locator[5:]
            el = self.page.locator(f"text={name}").first
        else:
            el = self.page.locator(locator).first
        el.wait_for(state="visible", timeout=10000)
        # 原生 select 优先按 label/value 选择。
        try:
            el.select_option(label=option)
            return
        except Exception:
            pass

        # 自定义下拉（如 Ant Design）：展开后按文本选择，最后才退到第一项。
        el.click()
        self.page.wait_for_timeout(500)
        if option:
            for candidate in (
                self.page.get_by_text(option, exact=True).first,
                self.page.locator(f".ant-select-item-option:has-text('{option}')").first,
            ):
                try:
                    if candidate.is_visible(timeout=1500):
                        candidate.click()
                        return
                except Exception:
                    continue
        fallback = self.page.locator('.ant-select-item-option').first
        if fallback.is_visible(timeout=2000):
            fallback.click()
            return
        raise ValueError(f"找不到下拉选项: {option}")

    def _verify(self, locator: str):
        if not locator:
            raise ValueError("verify 操作缺少元素定位器")
        self.page.locator(locator).first.wait_for(state="visible", timeout=10000)

    def _hover(self, locator: str):
        if not locator:
            raise ValueError("hover 操作缺少元素定位器")
        self.page.locator(locator).first.hover(timeout=10000)
    def _scroll(self, locator: str):
        self.page.evaluate("window.scrollBy(0, 600)")


# ═══════════════════════════════════════════════
#  AssertionExecutor — 断言执行
# ═══════════════════════════════════════════════

class AssertionExecutor:
    def __init__(self, page: Page):
        self.page = page

    def assert_by_type(self, assert_type: str, verify_point: str, last_loc: str = ""):
        if not assert_type or not verify_point:
            return

        logger.info(f"  断言: [{assert_type}] {verify_point}")

        if assert_type == "url_contains":
            self._url_contains(verify_point)
        elif assert_type == "url_not_contains":
            self._url_not_contains(verify_point)
        elif assert_type == "url_matches":
            self._url_matches(verify_point)
        elif assert_type == "text_contains":
            self._text_contains(verify_point, last_loc)
        elif assert_type == "text_contains_all":
            self._text_contains_all(verify_point, last_loc)
        elif assert_type == "text_equals":
            self._text_equals(verify_point, last_loc)
        elif assert_type in ("text_visible", "visible_text"):
            self._text_visible(verify_point)
        elif assert_type == "title_contains":
            self._title_contains(verify_point)
        elif assert_type == "element_visible":
            self._element_visible(verify_point)
        elif assert_type == "element_not_visible":
            self._element_not_visible(verify_point)
        else:
            raise ValueError(f"不支持的断言类型: {assert_type}")

    @staticmethod
    def _extract_keyword(text: str) -> str:
        m = re.search(r"'([^']+)'", text)
        return m.group(1).strip() if m else text[:40].strip()

    def _text_contains(self, expected: str, locator: str) -> bool:
        kw = self._extract_keyword(expected) or expected
        try: self.page.wait_for_load_state("domcontentloaded", timeout=5000)
        except Exception: pass
        if locator:
            try:
                if kw in self.page.locator(locator).first.inner_text(timeout=3000):
                    logger.info(f"  OK 元素含: {kw}"); return True
            except Exception: pass
        body = self.page.locator("body").inner_text()
        if kw in body:
            logger.info(f"  OK body含: {kw}"); return True
        raise AssertionError(f"文本不包含「{kw}」")

    def _text_contains_all(self, expected: str, locator: str) -> bool:
        values = [value.strip() for value in expected.split("|") if value.strip()]
        if not values:
            raise AssertionError("text_contains_all 缺少验证文本")
        for value in values:
            self._text_contains(value, locator)
        return True

    def _text_equals(self, expected: str, locator: str) -> bool:
        actual = self.page.locator(locator).first.inner_text() if locator else self.page.locator("body").inner_text()
        assert actual.strip() == expected.strip(), f"文本不等: 期望={expected}, 实际={actual[:80]}"
        return True

    def _text_visible(self, expected: str) -> bool:
        expected = self._extract_keyword(expected) or expected
        for _ in range(10):
            body = self.page.locator("body").inner_text()
            if expected in body: logger.info(f"  OK 文本可见"); return True
            self.page.wait_for_timeout(500)
        raise AssertionError(f"文本不可见: {expected}")

    def _element_visible(self, locator: str) -> bool:
        self.page.locator(locator).first.wait_for(state="visible", timeout=5000)
        logger.info(f"  OK 元素可见"); return True

    def _url_contains(self, expected: str) -> bool:
        kw = self._extract_keyword(expected) or expected
        assert kw in self.page.url, f"URL不含「{kw}」, 当前: {self.page.url}"
        logger.info(f"  OK URL含: {kw}"); return True

    def _url_not_contains(self, expected: str) -> bool:
        kw = self._extract_keyword(expected) or expected
        assert kw not in self.page.url, f"URL仍含「{kw}」, 当前: {self.page.url}"
        logger.info(f"  OK URL不含: {kw}")
        return True

    def _url_matches(self, pattern: str) -> bool:
        assert re.search(pattern, self.page.url), f"URL不匹配: {pattern}"
        return True

    def _title_contains(self, expected: str) -> bool:
        kw = self._extract_keyword(expected) or expected
        actual = self.page.title()
        assert kw in actual, f"标题不含「{kw}」, 当前: {actual}"
        logger.info(f"  OK 标题含: {kw}")
        return True

    def _element_not_visible(self, locator: str) -> bool:
        try:
            self.page.locator(locator).first.wait_for(state="visible", timeout=3000)
            raise AssertionError(f"元素仍可见: {locator}")
        except PwTimeout:
            logger.info(f"  OK 元素不可见")
            return True


# ═══════════════════════════════════════════════
#  GenericTestRunner — 测试框架入口
# ═══════════════════════════════════════════════

class GenericTestRunner:
    def __init__(self):
        self.store = StandardCaseStore()
        self.results = []
        # 环境变量中的登录凭证
        self.login_url = settings.LOGIN_URL
        self.login_username = settings.LOGIN_USERNAME
        self.login_password = settings.LOGIN_PASSWORD

    def run_all(self, headless: bool = False, case_filter: str = "",
                module_filter: str = ""):
        cases = self.store.load_cases()
        if module_filter:
            cases = [c for c in cases if c.get("模块", "") == module_filter]
        if case_filter:
            cases = [c for c in cases if c["用例ID"] == case_filter]
            if not cases:
                logger.error(f"未找到用例: {case_filter}")
                return

        logger.info(f"共 {len(cases)} 条用例")

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless, args=["--disable-blink-features=AutomationControlled"])
            context = browser.new_context(no_viewport=True)
            page = context.new_page()

            login_url = settings.LOGIN_URL

            for case in cases:
                case_id = case["用例ID"]
                logger.info(f"\n{'='*50}\n  执行: [{case_id}] {case.get('测试场景','')}\n{'='*50}")

                start_t = time.time()
                result = "pass"
                error_msg = ""

                try:
                    self._run_one(page, case, login_url)
                except Exception as e:
                    result = f"fail: {str(e)[:200]}"
                    error_msg = str(e)
                    logger.error(f"  [{case_id}] 执行异常: {e}")

                duration = time.time() - start_t
                self.results.append({
                    "case_id": case_id,
                    "result": result,
                    "row_num": case.get("_row"),
                    "duration": round(duration, 1),
                })

                # 实时回写
                try:
                    row_num = case.get("_row")
                    if row_num:
                        store = StandardCaseStore()
                        wb = load_workbook(store.filepath)
                        ws = wb.active
                        cell = ws.cell(row=row_num, column=17)
                        cell.value = result
                        cell.font = Font(color="008000" if result == "pass" else "FF0000")
                        wb.save(store.filepath)
                        wb.close()
                except Exception as write_err:
                    logger.warning(f"  回写失败: {write_err}")

            self._print_summary()

    def _run_one(self, page: Page, case: dict, login_url: str):
        preconditions = case.get("前置条件", "")
        locators_str = case.get("元素定位器", "")
        operations_str = case.get("操作类型", "")
        data_str = case.get("输入数据", "")
        assert_type = case.get("断言类型", "")
        verify_point = case.get("验证点", "")
        expected = case.get("期望结果", "")

        # 1. 前置条件
        self._handle_preconditions(page, preconditions, login_url)

        # 2. 执行操作
        executor = StepExecutor(page)
        executor.execute(locators_str, operations_str, data_str)

        # 3. 断言
        loc_list = [locator for locator in _split_steps(locators_str) if locator]
        last_loc = loc_list[-1] if loc_list else ""
        page.wait_for_timeout(500)
        assertion = AssertionExecutor(page)
        assertion.assert_by_type(assert_type, verify_point or expected, last_loc)

    def _handle_preconditions(self, page: Page, preconditions: str, login_url: str):
        mode = _precondition_mode(preconditions)
        if mode == "none":
            return

        if mode == "open_login":
            if not login_url:
                raise ValueError("前置条件要求打开登录页，但未配置 LOGIN_URL")
            page.goto(login_url, wait_until="domcontentloaded", timeout=30000)

        elif mode == "auto_login":
            if not login_url:
                raise ValueError("前置条件要求自动登录，但未配置 LOGIN_URL")
            if not settings.LOGIN_USERNAME or not settings.LOGIN_PASSWORD:
                raise ValueError("前置条件要求自动登录，但未配置 LOGIN_USERNAME/LOGIN_PASSWORD")
            logger.info("  前置: 自动登录...")
            page.goto(login_url, wait_until="domcontentloaded", timeout=30000)
            page.wait_for_timeout(1000)

            uname = settings.LOGIN_USERNAME
            pwd = settings.LOGIN_PASSWORD
            # 输入账号
            try:
                page.locator("input[type='text']").first.fill(uname, timeout=5000)
            except Exception:
                try:
                    page.get_by_role("textbox").first.fill(uname, timeout=5000)
                except Exception:
                    page.locator("input").first.fill(uname, timeout=5000)
            # 输入密码
            try:
                page.locator("input[type='password']").first.fill(pwd, timeout=5000)
            except Exception:
                page.get_by_role("textbox").nth(1).fill(pwd, timeout=5000)

            # 选择门店
            try:
                store_sel = page.locator(".ant-select-selector").first
                store_sel.click()
                page.wait_for_timeout(600)
                opt = page.locator(".ant-select-item-option").first
                if opt.is_visible(timeout=2000):
                    opt.click()
            except Exception:
                pass

            # 点击登录
            login_btn = page.locator("button[type='submit']")
            if login_btn.count() == 0:
                login_btn = page.locator("button:has-text('登')")
            login_btn.first.click(timeout=5000)
            # 等待页面跳转离开登录页
            try:
                page.wait_for_url("**/login", timeout=3000)
            except Exception:
                pass
            try:
                page.wait_for_function("!window.location.href.includes('/login')", timeout=15000)
            except Exception:
                pass
            page.wait_for_timeout(2000)
            logger.info(f"  登录完成, URL={page.url}")

            # 关闭"知道了"弹窗
            try:
                btn = page.get_by_text("知道了").first
                if btn.is_visible(timeout=1000):
                    btn.click()
            except Exception:
                pass

        else:
            if "about:blank" in page.url or not page.url.startswith("http"):
                page.goto(login_url, wait_until="domcontentloaded", timeout=30000)
                logger.info(f"  前置: 导航到 {login_url}")

    def _print_summary(self):
        passed = sum(1 for r in self.results if r["result"] == "pass")
        total = len(self.results)
        print(f"\n{'='*50}")
        print(f"  执行汇总: {passed}/{total} 通过")
        for r in self.results:
            s = "PASS" if r["result"] == "pass" else "FAIL"
            print(f"  [{s}] {r['case_id']} ({r['duration']}s)")
        print(f"{'='*50}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--case", help="只执行指定用例ID")
    parser.add_argument("--headless", action="store_true")
    args = parser.parse_args()

    runner = GenericTestRunner()
    runner.run_all(headless=args.headless, case_filter=args.case)
