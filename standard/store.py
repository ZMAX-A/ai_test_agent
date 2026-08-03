"""
标准用例库 Excel 存储层

采用 17 列格式（16 列用例定义 + 1 列实际结果）：
  Sheet: "自动化测试用例"
  列: 用例ID | 模块 | 测试场景 | 测试点 | 优先级 | 前置条件 |
      操作步骤 | 元素定位器 | 操作类型 | 输入数据 | 数据类型 |
      期望结果 | 验证点 | 断言类型 | 超时(秒) | 备注 | 实际结果
"""

import os
import json
import hashlib
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None

# ── 默认路径 ──
DEFAULT_STORE_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "test_cases", "standard.xlsx"
)

# ── Sheet 名 ──
SHEET_NAME = "自动化测试用例"

# ═══════════════════════════════════════════════
#  17 列表头（16 列用例定义 + 实际结果）
# ═══════════════════════════════════════════════
CASE_HEADERS = [
    "用例ID",      # 0  - 唯一编号，如 TC-LOGIN-001
    "模块",        # 1  - 所属功能模块
    "测试场景",    # 2  - 场景描述
    "测试点",      # 3  - 验证目标
    "优先级",      # 4  - P0/P1/P2/P3
    "前置条件",    # 5  - "已登录" / "打开登录页" / ""
    "操作步骤",    # 6  - 给人看的步骤说明
    "元素定位器",  # 7  - CSS选择器，多步用逗号分隔
    "操作类型",    # 8  - click/input/select/verify/hover/scroll/nav/wait，逗号分隔
    "输入数据",    # 9  - 多值用 | 分隔
    "数据类型",    # 10 - 预留
    "期望结果",    # 11 - 断言描述
    "验证点",      # 12 - 断言关键词（引号包裹）
    "断言类型",    # 13 - text_contains/url_contains/element_visible...
    "超时(秒)",    # 14 - 默认5
    "备注",        # 15 - 补充说明
    "实际结果",    # 16 - ⛔ 引擎自动回写，人工不填
]

# ── 样式 ──
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN = Side(style="thin", color="999999")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)
WRAP_ALIGN = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


class StandardCaseStore:
    """标准用例库 Excel 存储 — 17列格式（含实际结果）"""

    def __init__(self, filepath: str = None):
        if openpyxl is None:
            raise ImportError("需要 openpyxl 库: pip install openpyxl")
        self.filepath = filepath or DEFAULT_STORE_PATH
        self._ensure_file()

    # ═══════════════════════════════════════════════
    #  文件初始化
    # ═══════════════════════════════════════════════

    def _ensure_file(self):
        if os.path.isfile(self.filepath):
            return
        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        for col_idx, h in enumerate(CASE_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = BORDER

        # 设置列宽
        widths = [16, 12, 22, 30, 8, 14, 30, 40, 20, 25, 10, 30, 25, 16, 10, 20, 25]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(65 + i) if i < 26 else 'A'].width = w

        wb.save(self.filepath)
        print(f"[STORE] 已初始化标准用例库(17列格式): {self.filepath}")

    @staticmethod
    def _style_cell(ws, row, col, value, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = BORDER
        cell.alignment = WRAP_ALIGN
        if fill:
            cell.fill = fill
        return cell

    # ═══════════════════════════════════════════════
    #  读取
    # ═══════════════════════════════════════════════

    def load_cases(self, module: str = "", priority: str = "",
                   only_enabled: bool = True) -> list[dict]:
        """加载全部用例"""
        cases = self._read_all_rows()
        result = []
        for c in cases:
            if module and c.get("模块", "") != module:
                continue
            if priority and c.get("优先级", "") != priority:
                continue
            result.append(c)
        return result

    def load_case(self, case_id: str) -> Optional[dict]:
        """加载单个用例"""
        for c in self._read_all_rows():
            if c["用例ID"] == case_id:
                return c
        return None

    # ═══════════════════════════════════════════════
    #  写入
    # ═══════════════════════════════════════════════

    def save_case(self, case: dict) -> str:
        """新增或更新一个用例，返回 case_id"""
        case_id = case.get("用例ID", case.get("case_id", ""))
        if not case_id:
            raise ValueError("用例ID 不能为空")

        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_sheet(wb)

        existing_row = self._find_row(ws, case_id)
        row_idx = existing_row or (ws.max_row + 1)

        # 按 CASE_HEADERS 顺序严格取值
        row_data = [
            case_id,                                                          # 0  用例ID
            case.get("模块", "") or case.get("module", ""),                   # 1  模块
            case.get("测试场景", "") or case.get("测试名称", "") or case.get("case_name", "") or case.get("name", ""),  # 2 测试场景
            case.get("测试点", "") or case.get("description", ""),            # 3  测试点
            case.get("优先级", "") or case.get("priority", "P1"),             # 4  优先级
            case.get("前置条件", "") or case.get("preconditions", ""),        # 5  前置条件
            case.get("操作步骤", "") or case.get("步骤描述", ""),             # 6  操作步骤
            case.get("元素定位器", "") or case.get("locators", ""),           # 7  元素定位器
            case.get("操作类型", "") or case.get("operations", ""),           # 8  操作类型
            case.get("输入数据", "") or case.get("input_data", ""),           # 9  输入数据
            case.get("数据类型", ""),                                          # 10 数据类型
            case.get("期望结果", "") or case.get("预期结果", "") or case.get("expected", ""),  # 11 期望结果
            case.get("验证点", "") or case.get("verify_point", ""),           # 12 验证点
            case.get("断言类型", "") or case.get("assert_type", ""),           # 13 断言类型
            str(case.get("超时(秒)", "") or case.get("timeout", "5")),        # 14 超时(秒)
            case.get("备注", "") or case.get("remark", ""),                   # 15 备注
            case.get("实际结果", "") or case.get("actual_result", ""),         # 16 实际结果
        ]

        for col_idx, val in enumerate(row_data, 1):
            self._style_cell(ws, row_idx, col_idx, val)

        wb.save(self.filepath)
        wb.close()
        return case_id

    def delete_case(self, case_id: str) -> bool:
        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_sheet(wb)
        row = self._find_row(ws, case_id)
        if row:
            ws.delete_rows(row)
            wb.save(self.filepath)
            wb.close()
            return True
        wb.close()
        return False

    # ═══════════════════════════════════════════════
    #  结果回写
    # ═══════════════════════════════════════════════

    def write_result(self, case_id: str, result: str, row_num: int = None):
        """回写单条结果到「实际结果」列"""
        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_sheet(wb)

        # 「实际结果」是第17列（索引17）
        result_col = 17

        if row_num:
            fill = PASS_FILL if result.startswith("pass") else FAIL_FILL
            self._style_cell(ws, row_num, result_col, result, fill=fill)
        else:
            row = self._find_row(ws, case_id)
            if row:
                fill = PASS_FILL if result.startswith("pass") else FAIL_FILL
                self._style_cell(ws, row, result_col, result, fill=fill)

        wb.save(self.filepath)
        wb.close()

    def write_results_batch(self, results: list[dict]):
        """批量回写结果 [{case_id, result, row_num}]"""
        wb = openpyxl.load_workbook(self.filepath)
        ws = self._get_sheet(wb)
        result_col = 17

        for r in results:
            case_id = r.get("case_id", "")
            result = r.get("result", "")
            row_num = r.get("row_num")
            if row_num:
                fill = PASS_FILL if "pass" in str(result).lower() else FAIL_FILL
                self._style_cell(ws, row_num, result_col, str(result), fill=fill)
            else:
                row = self._find_row(ws, case_id)
                if row:
                    fill = PASS_FILL if "pass" in str(result).lower() else FAIL_FILL
                    self._style_cell(ws, row, result_col, str(result), fill=fill)

        wb.save(self.filepath)
        wb.close()

    # ═══════════════════════════════════════════════
    #  变更检测
    # ═══════════════════════════════════════════════

    def detect_changes(self, manifest_path: str = None) -> dict:
        if manifest_path is None:
            manifest_path = os.path.join(
                os.path.dirname(os.path.dirname(self.filepath)),
                "generated_scripts", ".generation_manifest.json"
            )

        cases = self._read_all_rows()
        current = {}
        for c in cases:
            content = json.dumps(c, ensure_ascii=False, sort_keys=True)
            current[c["用例ID"]] = hashlib.sha256(content.encode()).hexdigest()[:16]

        previous = {}
        if os.path.isfile(manifest_path):
            try:
                with open(manifest_path, "r", encoding="utf-8") as f:
                    previous = json.load(f)
            except Exception:
                pass

        changed, new_cases, unchanged, deleted = [], [], [], []
        for cid, h in current.items():
            if cid not in previous:
                new_cases.append(cid)
            elif previous[cid] != h:
                changed.append(cid)
            else:
                unchanged.append(cid)
        for cid in previous:
            if cid not in current:
                deleted.append(cid)

        return {
            "changed": changed, "new": new_cases,
            "deleted": deleted, "unchanged": unchanged,
            "total": len(current),
        }

    def save_manifest(self, manifest_path: str = None):
        if manifest_path is None:
            manifest_path = os.path.join(
                os.path.dirname(os.path.dirname(self.filepath)),
                "generated_scripts", ".generation_manifest.json"
            )
        os.makedirs(os.path.dirname(manifest_path), exist_ok=True)
        cases = self._read_all_rows()
        manifest = {}
        for c in cases:
            content = json.dumps(c, ensure_ascii=False, sort_keys=True)
            manifest[c["用例ID"]] = hashlib.sha256(content.encode()).hexdigest()[:16]
        with open(manifest_path, "w", encoding="utf-8") as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)

    # ═══════════════════════════════════════════════
    #  统计
    # ═══════════════════════════════════════════════

    def get_stats(self) -> dict:
        cases = self._read_all_rows()
        by_status = {}
        by_module = {}
        for c in cases:
            m = c.get("模块", "__root__")
            by_module[m] = by_module.get(m, 0) + 1
            r = c.get("实际结果", "")
            s = "pass" if r and "pass" in str(r).lower() else ("fail" if r else "untested")
            by_status[s] = by_status.get(s, 0) + 1
        return {"total": len(cases), "by_status": by_status, "by_module": by_module}

    # ═══════════════════════════════════════════════
    #  内部
    # ═══════════════════════════════════════════════

    def _get_sheet(self, wb):
        if SHEET_NAME in wb.sheetnames:
            return wb[SHEET_NAME]
        # 兼容旧版 sheet 名
        ws = wb.active
        if ws.title != SHEET_NAME:
            ws.title = SHEET_NAME
        return ws

    def _read_all_rows(self) -> list[dict]:
        if not os.path.isfile(self.filepath):
            return []
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        ws = self._get_sheet(wb)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        headers = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        wb.close()

        cases = []
        for row_idx, row in enumerate(rows, 2):
            if all(v is None for v in row):
                continue
            d = dict(zip(headers, row))
            cid = str(d.get("用例ID") or "").strip()
            if not cid:
                continue
            case = {}
            for i, h in enumerate(CASE_HEADERS):
                val = row[i] if i < len(row) else ""
                case[h] = str(val).strip() if val is not None else ""
            case["_row"] = row_idx
            cases.append(case)
        return cases

    @staticmethod
    def _find_row(ws, case_id: str) -> int:
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row=row, column=1).value or "").strip()
            if val == case_id:
                return row
        return 0


# ── 便捷单例 ──
def get_store(filepath: str = None) -> StandardCaseStore:
    if not hasattr(get_store, "_instance"):
        get_store._instance = StandardCaseStore(filepath)
    elif filepath and filepath != get_store._instance.filepath:
        get_store._instance = StandardCaseStore(filepath)
    return get_store._instance


if __name__ == "__main__":
    store = StandardCaseStore()
    stats = store.get_stats()
    print(f"标准用例库: {store.filepath}")
    print(f"  用例总数: {stats['total']}")
    print(f"  按状态: {stats['by_status']}")
    cases = store.load_cases()
    for c in cases:
        print(f"  {c['用例ID']:16s} [{c['优先级']}] {c['测试场景'][:30]}")
