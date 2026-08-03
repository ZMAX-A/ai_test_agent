"""
标准用例 Excel 导出工具

将 case_library/ 中的标准用例（JSON）导出为 Excel 文件，
包含用例清单和步骤详情两个 Sheet，方便人工审核、打印、归档。

用法：
    python scripts/export_standard_cases.py                          # 导出全部用例
    python scripts/export_standard_cases.py --module 搜索功能        # 按模块导出
    python scripts/export_standard_cases.py --output report.xlsx     # 指定输出文件
    python scripts/export_standard_cases.py --status draft           # 按状态筛选
"""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__))))
sys.stdout.reconfigure(encoding="utf-8")

from case import case_manager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT_DEFAULT = "test_cases/standard_cases.xlsx"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin = Side(style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
center_align = Alignment(horizontal="center", vertical="center")
wrap_align = Alignment(vertical="center", wrap_text=True)


def _style_header(ws, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border


def _style_data(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = wrap_align


def _set_col_widths(ws, widths):
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[letters[i - 1]].width = w


def export_to_excel(module: str = "", status: str = "", output: str = OUTPUT_DEFAULT):
    """将标准用例导出为 Excel"""
    # ── 获取用例 ──
    if module and status:
        cases = [c for c in case_manager.find_cases(module=module) if c.get("status") == status]
    elif module:
        cases = case_manager.find_cases(module=module)
    elif status:
        cases = case_manager.find_cases(status=status)
    else:
        cases = case_manager.find_cases()

    if not cases:
        filters = []
        if module:
            filters.append(f"模块={module}")
        if status:
            filters.append(f"状态={status}")
        tag = "（" + ", ".join(filters) + "）" if filters else ""
        print(f"[!] 无标准用例可导出{tag}")
        return

    # 排序：模块 → case_id
    cases.sort(key=lambda c: (c.get("module", ""), c.get("case_id", "")))

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════
    # Sheet 1: 用例清单
    # ════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "用例清单"

    headers1 = [
        "用例编号", "用例名称", "模块", "版本", "状态",
        "前置条件", "起始网址", "步骤数", "来源任务",
        "创建时间", "更新时间",
    ]
    _style_header(ws1, headers1)

    row_idx = 2
    for case in cases:
        steps = case.get("steps", [])
        ws1.append([
            case.get("case_id", ""),
            case.get("name", ""),
            case.get("module", ""),
            case.get("version", 1),
            _status_label(case.get("status", "")),
            case.get("preconditions", ""),
            case.get("start_url", ""),
            len(steps),
            case.get("source_task", ""),
            case.get("created_at", ""),
            case.get("updated_at", ""),
        ])
        row_idx += 1

    _style_data(ws1, 2, row_idx - 1, len(headers1))
    _set_col_widths(ws1, [12, 24, 14, 8, 14, 24, 36, 10, 20, 20, 20])
    ws1.freeze_panes = "A2"

    # ════════════════════════════════════════════════════
    # Sheet 2: 步骤详情
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="步骤详情")

    headers2 = [
        "用例编号", "用例名称", "模块", "步骤号", "步骤目标",
        "动作类型", "动作参数(JSON)", "断言数", "断言详情",
    ]
    _style_header(ws2, headers2)

    row_idx = 2
    for case in cases:
        steps = case.get("steps", [])
        case_id = case.get("case_id", "")
        case_name = case.get("name", "")
        module_name = case.get("module", "")

        for step in steps:
            step_num = step.get("step", 0)
            goal = step.get("goal", "")
            action = step.get("action", "")
            params = step.get("parameters", {})
            asserts = step.get("asserts", [])

            # 参数精简展示：去掉空值
            clean_params = {k: v for k, v in params.items() if v not in (None, "", [])}
            params_str = json_dumps(clean_params)

            # 断言详情
            assert_details = "; ".join(
                [f"{a.get('type','')}: {a.get('target','')}" for a in asserts]
            ) if asserts else ""

            ws2.append([
                case_id,
                case_name,
                module_name,
                step_num,
                goal,
                action,
                params_str,
                len(asserts),
                assert_details,
            ])
            row_idx += 1

    _style_data(ws2, 2, row_idx - 1, len(headers2))
    _set_col_widths(ws2, [12, 24, 14, 10, 36, 14, 50, 10, 50])
    ws2.freeze_panes = "A2"

    # ════════════════════════════════════════════════════
    # Sheet 3: 断言明细
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet(title="断言明细")

    headers3 = [
        "用例编号", "步骤号", "步骤目标", "断言类型", "断言目标", "断言说明",
    ]
    _style_header(ws3, headers3)

    row_idx = 2
    for case in cases:
        steps = case.get("steps", [])
        case_id = case.get("case_id", "")

        for step in steps:
            step_num = step.get("step", 0)
            goal = step.get("goal", "")
            asserts = step.get("asserts", [])

            for a in asserts:
                ws3.append([
                    case_id,
                    step_num,
                    goal,
                    a.get("type", ""),
                    a.get("target", ""),
                    _assert_description(a),
                ])
                row_idx += 1

    _style_data(ws3, 2, row_idx - 1, len(headers3))
    _set_col_widths(ws3, [12, 10, 36, 18, 40, 40])
    ws3.freeze_panes = "A2"

    # ── 保存 ──
    wb.save(output)
    status_tag = f"（状态={status}）" if status else ""
    module_tag = f"（模块={module}）" if module else ""
    print(f"[OK] 已导出 {len(cases)} 个标准用例: {output}")
    print(f"      模块: {module_tag or '全部'} {status_tag}")
    print(f"      Sheet1「用例清单」: {len(cases)} 行")
    print(f"      Sheet2「步骤详情」: 含所有步骤及参数")
    print(f"      Sheet3「断言明细」: 所有断言独立行")


# ── 辅助函数 ──

def json_dumps(obj) -> str:
    """安全 JSON 序列化（单行）"""
    import json
    try:
        return json.dumps(obj, ensure_ascii=False, separators=(",", ": "))
    except Exception:
        return str(obj)


_STATUS_LABELS = {
    "active": "激活",
    "draft": "草稿",
    "deprecated": "已废弃",
    "needs_update": "待更新",
}


def _status_label(s: str) -> str:
    return _STATUS_LABELS.get(s, s)


_ASSERT_DESC = {
    "url_contains": "当前 URL 包含指定文本",
    "url_exact": "当前 URL 完全匹配",
    "title_contains": "页面标题包含指定文本",
    "text_exists": "页面上存在指定文本",
    "element_visible": "指定元素可见",
}


def _assert_description(a: dict) -> str:
    base = _ASSERT_DESC.get(a.get("type", ""), "")
    return base


if __name__ == "__main__":
    # ── 解析参数 ──
    module = ""
    status = ""
    output = OUTPUT_DEFAULT

    args = sys.argv[1:]
    i = 0
    while i < len(args):
        if args[i] == "--module" and i + 1 < len(args):
            module = args[i + 1]
            i += 2
        elif args[i] == "--status" and i + 1 < len(args):
            status = args[i + 1]
            i += 2
        elif args[i] == "--output" and i + 1 < len(args):
            output = args[i + 1]
            i += 2
        else:
            i += 1

    export_to_excel(module=module, status=status, output=output)
