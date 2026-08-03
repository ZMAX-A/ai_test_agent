"""
生成 Excel 测试用例模板：用例清单 + 步骤明细两个 Sheet
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = "test_cases/test_cases.xlsx"

wb = openpyxl.Workbook()

# ── 通用样式 ──
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin = Side(style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
center_align = Alignment(horizontal="center", vertical="center")
wrap_align = Alignment(vertical="center", wrap_text=True)


def style_header(ws, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border


def style_data(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = wrap_align


def set_col_widths(ws, widths):
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[letters[i - 1]].width = w


# ═══════════════════════════════════════════
# Sheet 1: 用例清单
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "用例清单"

headers1 = [
    "case_id", "case_name", "module", "mode",
    "start_url", "task", "priority", "tags",
    "enabled", "description", "login_required", "credential_key",
]
style_header(ws1, headers1)

# 示例数据
sample1 = [
    ["TC001", "百度搜索世界杯", "搜索功能", "auto",
     "https://www.baidu.com",
     '打开百度，搜索"世界杯"，验证搜索结果中包含"世界杯"',
     "P1", "回归,搜索", "Y", "验证百度搜索基本功能正常", "N", ""],
    ["TC002", "百度登录页导航", "登录功能", "auto",
     "https://www.baidu.com",
     "点击登录，验证跳转到登录页面",
     "P1", "登录", "Y", "", "N", ""],
    ["TC003", "12306搜索", "搜索功能", "steps",
     "https://www.baidu.com",
     "",
     "P2", "", "Y", "结构化步骤示例，步骤在步骤明细Sheet中定义", "N", ""],
    ["TC010", "顾客档案查询", "业务功能", "steps",
     "{{LOGIN_URL}}",
     "",
     "P0", "冒烟,业务", "Y", "login_required自动注入登录步骤，此处只需写登录后的操作", "Y", "prod"],
]
for row in sample1:
    ws1.append(row)

style_data(ws1, 2, 1 + len(sample1), len(headers1))
set_col_widths(ws1, [10, 22, 12, 10, 30, 60, 10, 18, 10, 35, 14, 16])
ws1.freeze_panes = "A2"

# 数据验证：mode 列
dv_mode = DataValidation(type="list", formula1='"auto,steps"', allow_blank=True)
dv_mode.sqref = "D2:D201"
dv_mode.prompt = "选择运行模式"
dv_mode.promptTitle = "模式"
ws1.add_data_validation(dv_mode)

# 数据验证：enabled 列
dv_enabled = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv_enabled.sqref = "I2:I201"
dv_enabled.prompt = "Y=启用  N=禁用"
dv_enabled.promptTitle = "启用状态"
ws1.add_data_validation(dv_enabled)

# 数据验证：priority 列
dv_pri = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
dv_pri.sqref = "G2:G201"
dv_pri.prompt = "P0=最高  P3=最低"
dv_pri.promptTitle = "优先级"
ws1.add_data_validation(dv_pri)

# 数据验证：login_required 列
dv_login = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
dv_login.sqref = "K2:K201"
dv_login.prompt = "Y=需预先登录 N=不需要"
dv_login.promptTitle = "需要登录"
ws1.add_data_validation(dv_login)

# ═══════════════════════════════════════════
# Sheet 2: 步骤明细
# ═══════════════════════════════════════════
ws2 = wb.create_sheet(title="步骤明细")

headers2 = [
    "case_id", "step_no", "goal",
    "assert_type", "assert_target", "assert_value", "remark",
]
style_header(ws2, headers2)

sample2 = [
    ["TC003", 1, "打开百度首页", "url_contains", "baidu", "", ""],
    ["TC003", 2, "在搜索框输入12306", "", "", "", ""],
    ["TC003", 3, "点击百度一下按钮", "text_exists", "12306", "", "验证搜索结果出现"],
    ["TC001", 1, "搜索世界杯", "text_exists", "世界杯", "", "auto模式可以不由步骤明细驱动"],
    ["TC010", 1, "点击菜单-顾客档案", "", "", "", "login_required注入的登录步骤执行完成后，开始执行这里的步骤"],
    ["TC010", 2, "验证顾客档案页面加载", "title_contains", "顾客档案", "", ""],
]
for row in sample2:
    ws2.append(row)

style_data(ws2, 2, 1 + len(sample2), len(headers2))
set_col_widths(ws2, [10, 10, 35, 18, 30, 30, 35])
ws2.freeze_panes = "A2"

# 数据验证：assert_type 列
dv_at = DataValidation(
    type="list",
    formula1='"url_contains,url_exact,title_contains,text_exists,element_visible"',
    allow_blank=True,
)
dv_at.sqref = "D2:D501"
dv_at.prompt = "选择断言类型"
dv_at.promptTitle = "断言类型"
ws2.add_data_validation(dv_at)

# ── 保存 ──
wb.save(OUTPUT)
print(f"[OK] Excel 模板已生成: {OUTPUT}")
print(f"      Sheet1「用例清单」: {len(headers1)} 列")
print(f"      Sheet2「步骤明细」: {len(headers2)} 列")
print()
print("打开 Excel 即可看到带表头、示例数据和下拉菜单的模板。")
