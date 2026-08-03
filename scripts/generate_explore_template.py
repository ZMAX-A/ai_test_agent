
"""
生成探索模式 Excel 模板

列：用例编号、用例名称、模块、前置条件、步骤描述、预期结果、起始网址
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = "test_cases/explore_cases.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "探索用例"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin = Side(style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
wrap_align = Alignment(vertical="center", wrap_text=True)

headers = [
    "用例编号", "用例名称", "模块", "前置条件",
    "起始网址", "步骤描述", "预期结果",
]

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

samples = [
    [
        "TC001", "顾客档案查询", "业务功能",
        "需配置 .env 登录凭证",
        "{{LOGIN_URL}}",
        "打开登录页面 {{LOGIN_URL}}\n"
        "在用户名输入框(第1个)输入 {{LOGIN_USERNAME}}\n"
        "在密码输入框(第2个)输入 {{LOGIN_PASSWORD}}\n"
        "点击登录按钮\n进入顾客档案页面(可直接访问/customer路径)",
        "成功进入顾客档案管理页面",
    ],
]

for row_idx, sample in enumerate(samples, 2):
    for col_idx, val in enumerate(sample, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = border
        cell.alignment = wrap_align

widths = [12, 22, 14, 20, 35, 50, 35]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

ws.freeze_panes = "A2"
wb.save(OUTPUT)
print(f"[OK] 探索用例模板已生成: {OUTPUT}")
print(f"     列: {' | '.join(headers)}")
print()
print("提示：")
print("  - 步骤描述中的每一行 | 分隔为一步")
print("  - {{VAR}} 会自动替换为 .env 中的值")
print("  - 起始网址为空时默认使用 https://www.baidu.com")
